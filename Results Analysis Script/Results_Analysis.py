#**************************************************************************** Evaluation Metrics Calculation  ****************************************************************************

import sys
import os
import fnmatch
import glob, os
import subprocess
import math
import numpy as np
import time
import statistics
from openpyxl import Workbook, load_workbook
import openpyxl




#**************************************************************************** Open Microservices Data Stored in Knowledge Base  ****************************************************************************

def open_files(scenario, filename, run):

    if run == 1:
        workbook = load_workbook(f'./Test/{scenario}/Run 1/{filename}.xlsx')
    elif run == 2:
        workbook = load_workbook(f'./Test/{scenario}/Run 2/{filename}.xlsx')
    elif run == 3:
        workbook = load_workbook(f'./Test/{scenario}/Run 3/{filename}.xlsx')
    elif run == 4:
        workbook = load_workbook(f'./Test/{scenario}/Run 4/{filename}.xlsx')
    elif run == 5:
        workbook = load_workbook(f'./Test/{scenario}/Run 5/{filename}.xlsx')
    elif run == 6:
        workbook = load_workbook(f'./Test/{scenario}/Run 6/{filename}.xlsx')
    elif run == 7:
        workbook = load_workbook(f'./Test/{scenario}/Run 7/{filename}.xlsx')
    elif run == 8:
        workbook = load_workbook(f'./Test/{scenario}/Run 8/{filename}.xlsx')
    elif run == 9:
        workbook = load_workbook(f'./Test/{scenario}/Run 9/{filename}.xlsx')
    elif run == 10:
        workbook = load_workbook(f'./Test/{scenario}/Run 10/{filename}.xlsx')
    
    return workbook


#**************************************************************************** Main Code for stored data analysis Starts  ****************************************************************************



scenario = "20%"                                   #Starting data analysis from 1st scenario (i.e., CPU Thereshold 20% with 2 Max. Replicas)
replicas = 2       
filename = "frontend"                              #Starting data analysis from frontend microservice
CPU_Request = 100                                  #CPU Request value for frontend microservice
                                       
run = 1                                            #Starting from the first run analysis and continues uptil 10th run

# Selection of relevant scenario data file stored in Knowledge Base for data analysis

while (run <= 10):

    if scenario == "20%":
        CPU_Threshold = 20
        folder = str(replicas) + " Replicas + " + scenario
        workbook = open_files (folder, filename, run)

    
    elif scenario == "50%":
        CPU_Threshold = 50
        folder = str(replicas) + " Replicas + " + scenario
        workbook = open_files (folder, filename, run)

        
    elif scenario == "80%":
        CPU_Threshold = 80
        folder = str(replicas) + " Replicas + " + scenario
        workbook = open_files (folder, filename, run)

    
    
    
    
    
    worksheet = workbook.active


    test_time = []                                       #Getting load test time stamps from the stored data
    for cell in worksheet['A'][1:]:
        if cell.value is not None:
            test_time.append(cell.value)


    cpu_utilization = []                                 #Getting CPU Utlization (percentage) values
    for cell in worksheet['B'][1:]:
        if cell.value is not None:
            cpu_utilization.append(cell.value)
    Avg_CPU_Utilization = statistics.mean(cpu_utilization)



    #**************************************************************************** CPU Overutlization Calculation ********************************************************************


    
    overutilized_cpu = []                                 #Getting CPU Overutlization values where cpu_utilization > CPU_Threshold
    for i in range(len(cpu_utilization)):
        if cpu_utilization[i] > CPU_Threshold:
            overutilized_cpu.append(cpu_utilization[i])
        else:
            overutilized_cpu.append(0)


    Avg_overutilized_CPU = statistics.mean(overutilized_cpu)



    #**************************************************************************** Supply CPU Calculation ********************************************************************

    
    current_replicas = []
    for cell in worksheet['C'][1:]:
        if cell.value is not None:
            current_replicas.append(cell.value)



    Avg_used_allocated_cpu_list = [value * CPU_Request for value in current_replicas]

    Avg_used_allocated_cpu = statistics.mean(Avg_used_allocated_cpu_list)



    #**************************************************************************** Resource Capacity Calculation ******************************************************

    max_replicas = []
    for cell in worksheet['E'][1:]:
        if cell.value is not None:
            max_replicas.append(cell.value)



    Avg_max_available_cpu_list = [value * CPU_Request for value in max_replicas]

    Avg_max_available_cpu = statistics.mean (Avg_max_available_cpu_list)


    #**************************************************************************** Resource Demand Calculation ***********************************************************


    desired_replicas = []

    for cell in worksheet['D'][1:]:
        if cell.value is not None:
            desired_replicas.append(cell.value)


    Avg_desired_cpu_list = [value * CPU_Request for value in desired_replicas]

    Avg_desired_cpu = statistics.mean (Avg_desired_cpu_list)


    #**************************************************************************** Overprovisioned CPU Calculation **************************************************************

    Residual_replicas = []


    for i in range(len(max_replicas)):
        if max_replicas[i] > desired_replicas[i]:
            Residual_replicas.append(max_replicas[i] - desired_replicas[i])
        else:
            Residual_replicas.append(0)

    Avg_residual_cpu_list = [value * CPU_Request for value in Residual_replicas]

    Avg_residual_cpu = statistics.mean (Avg_residual_cpu_list)




    #**************************************************************************** Underprovisioned CPU Calculation ******************************************************


    required_replicas = []


    for i in range(len(max_replicas)):
        if max_replicas[i] < desired_replicas[i]:
            required_replicas.append(desired_replicas[i] - max_replicas[i])
        else:
            required_replicas.append(0)

    Avg_required_cpu_list = [value * CPU_Request for value in required_replicas]

    Avg_required_cpu = statistics.mean (Avg_required_cpu_list)

    


    #**************************************************************************** Overutilization Time Calculation *******************************************************************************

    Overutilization_Time_values = []
    Overutilization_Times = []
    Violation_start_time = 0
    Violation_end_time = 0
    start_index = 0
    end_index = 0

   

    for i in range(len(test_time)):
        Overutilization_Time_values[i] = test_time [i]


    for i in range(len(overutilized_cpu)):
        if overutilized_cpu[i] == 0:
            Overutilization_Time_values[i] = 0


    for i in range(len(Overutilization_Time_values)):
        
        if Overutilization_Time_values [i] > 0 and Overutilization_Time_values [i-1] == 0 and Violation_start_time == 0:
            Violation_start_time = Overutilization_Time_values [i]  
            start_index = i
        
        
        if Overutilization_Time_values [i] == 0 and Overutilization_Time_values [i-1] > 0  and Violation_start_time != 0:
            Violation_end_time = Overutilization_Time_values [i-1]
            end_index = i-1

        elif i == len(Overutilization_Time_values)-1 and Overutilization_Time_values[i] != 0 and Violation_start_time != 0:    #condition for test ending (like at 900sec)
            Violation_end_time = Overutilization_Time_values [i]
            end_index = i
                
        if Violation_start_time !=0  and Violation_end_time != 0:
            Overutilization_Times.append(Violation_end_time - Violation_start_time)
            Violation_start_time = 0
            Violation_end_time = 0

    Overutilization_Times_sum = sum (Overutilization_Times)   # Total Overutlization Time
    


    #**************************************************************************** Overprovisioned Time Calculation ******************************************************
 
    
    Overprovisioned_Time_values = []
    for i in range(len(test_time)):
        Overprovisioned_Time_values[i] = test_time [i]


    for i in range(len(Avg_residual_cpu_list)):
        if Avg_residual_cpu_list[i] == 0:
            Overprovisioned_Time_values[i] = 0

    
    Overprovisioned_Times = []
    Violation_start_time = 0
    Violation_end_time = 0
    start_index = 0
    end_index = 0



    for i in range(len(Overprovisioned_Time_values)):
        
        if Overprovisioned_Time_values [i] > 0 and Overprovisioned_Time_values [i-1] == 0 and Violation_start_time == 0:
            Violation_start_time = Overprovisioned_Time_values [i]  
            start_index = i
        
        
        if Overprovisioned_Time_values [i] == 0 and Overprovisioned_Time_values [i-1] > 0 and Violation_start_time != 0:
            Violation_end_time = Overprovisioned_Time_values [i-1]
            end_index = i-1

        elif i == len(Overprovisioned_Time_values)-1 and Overprovisioned_Time_values[i] != 0 and Violation_start_time != 0:                #condition for test ending (like at 900sec)
            Violation_end_time = Overprovisioned_Time_values [i]
            end_index = i
                
        if Violation_start_time !=0  and Violation_end_time != 0:
            Overprovisioned_Times.append(Violation_end_time - Violation_start_time)
            Violation_start_time = 0
            Violation_end_time = 0


    Overprovisioned_Times_sum = sum (Overprovisioned_Times)                   # Total Overprovisioned Time


   


    #**************************************************************************** Underprovisioned Time Calculation ******************************************************


    Underprovisioned_Time_values = []
    for i in range(len(test_time)):
        Underprovisioned_Time_values[i] = test_time [i]


    for i in range(len(Avg_required_cpu_list)):
        if Avg_required_cpu_list[i] == 0:
            Underprovisioned_Time_values[i] = 0
    

    Underprovisioned_Times = []
    Violation_start_time = 0
    Violation_end_time = 0
    start_index = 0
    end_index = 0


    for i in range(len(Underprovisioned_Time_values)):
        
        if Underprovisioned_Time_values [i] > 0 and Underprovisioned_Time_values [i-1] == 0 and Violation_start_time == 0:
            Violation_start_time = Underprovisioned_Time_values [i]  
            start_index = i
        
        
        if Underprovisioned_Time_values [i] == 0 and Underprovisioned_Time_values [i-1] > 0 and Violation_start_time != 0:
            Violation_end_time = Underprovisioned_Time_values [i-1]
            end_index = i-1

        elif i == len(Underprovisioned_Time_values)-1 and Underprovisioned_Time_values[i] != 0 and Violation_start_time != 0:                         #condition for test ending (like at 900sec)
            Violation_end_time = Underprovisioned_Time_values [i]
            end_index = i
                
        if Violation_start_time !=0  and Violation_end_time != 0:
            Underprovisioned_Times.append(Violation_end_time - Violation_start_time)
            Violation_start_time = 0
            Violation_end_time = 0



    Underprovisioned_Times_sum = sum (Underprovisioned_Times)          # Total Underprovisioned Time



  


    #********************************************************************* Storing Analyzed Results in a Seperate File ***********************************************************************

    workbook = load_workbook(f'./Knowledge Base/{filename}.xlsx')

    worksheet = workbook.active
    r= 0
    c= 0

    if scenario == "20%" and replicas == 2:
        r = 5
    elif scenario == "50%" and replicas == 2:
        r = 27
    elif scenario == "80%" and replicas == 2:
        r = 48
    if scenario == "20%" and replicas == 5:
        r = 72
    elif scenario == "50%" and replicas == 5:
        r = 93
    elif scenario == "80%" and replicas == 5:
        r = 114
    if scenario == "20%" and replicas == 10:
        r = 136
    elif scenario == "50%" and replicas == 10:
        r = 157
    elif scenario == "80%" and replicas == 10:
        r = 178


    if run == 1:
        c = 9
    elif run == 2:
        c = 11
    elif run == 3:
        c = 13
    elif run == 4:
        c = 15
    elif run == 5:
        c = 17
    elif run == 6:
        c = 19
    elif run == 7:
        c = 21
    elif run == 8:
        c = 23
    elif run == 9:
        c = 25
    elif run == 10:
        c = 27
    

    worksheet.cell(row=r, column=c, value=Avg_CPU_Utilization)
    r = r + 1
    worksheet.cell(row=r, column=c, value=Avg_overutilized_CPU)
    r = r + 1
    worksheet.cell(row=r, column=c, value=Overutilization_Times_sum)
    r = r + 1
    worksheet.cell(row=r, column=c, value=Avg_used_allocated_cpu)
    r = r + 1
    worksheet.cell(row=r, column=c, value=Avg_desired_cpu)
    r = r + 1
    worksheet.cell(row=r, column=c, value=Avg_max_available_cpu)
    r = r + 1
    worksheet.cell(row=r, column=c, value=Avg_residual_cpu)
    r = r + 1
    
    worksheet.cell(row=r, column=c, value=Overprovisioned_Times_sum)
    r = r + 1
    worksheet.cell(row=r, column=c, value=Avg_required_cpu)
    r = r + 1
    
    worksheet.cell(row=r, column=c, value=Underprovisioned_Times_sum)
    r = r + 1
    

    workbook.save(f'./Knowledge Base/{filename}.xlsx')

    


    # ************************************************* Switching to Next CPU threshold Scenario within same replica limit ***********************************************************************
    
    

    if run == 10 and scenario == "20%":
        run = 1
        scenario = "50%"
    
    elif run == 10 and scenario == "50%":
        run = 1
        scenario = "80%"



    # ********************************************************************* Switching to Next Microservice ***********************************************************************
    
    if run == 10 and scenario == "80%" and filename == "frontend":
        filename = "adservice"
        run = 1
        scenario = "20%"
        CPU_Request = 200
    
    elif run == 10 and scenario == "80%" and filename == "adservice":
        filename = "cartservice"
        run = 1
        scenario = "20%"
        CPU_Request = 200

    elif run == 10 and scenario == "80%" and filename == "cartservice":
        filename = "paymentservice"
        run = 1
        scenario = "20%"
        CPU_Request = 100

    elif run == 10 and scenario == "80%" and filename == "paymentservice":
        filename = "currencyservice"
        run = 1
        scenario = "20%"
        CPU_Request = 100

    elif run == 10 and scenario == "80%" and filename == "currencyservice":
        filename = "emailservice"
        run = 1
        scenario = "20%"
        CPU_Request = 100

    elif run == 10 and scenario == "80%" and filename == "emailservice":
        filename = "checkoutservice"
        run = 1
        scenario = "20%"
        CPU_Request = 100

    elif run == 10 and scenario == "80%" and filename == "checkoutservice":
        filename = "productcatalogservice"
        run = 1
        scenario = "20%"
        CPU_Request = 100
    
    elif run == 10 and scenario == "80%" and filename == "productcatalogservice":
        filename = "recommendationservice"
        run = 1
        scenario = "20%"
        CPU_Request = 100

    elif run == 10 and scenario == "80%" and filename == "recommendationservice":
        filename = "redis-cart"
        run = 1
        scenario = "20%"
        CPU_Request = 70

    elif run == 10 and scenario == "80%" and filename == "redis-cart":
        filename = "shippingservice"
        run = 1
        scenario = "20%"
        CPU_Request = 100

    

# ********************************************************************* Switching to Next Replica limit and CPU threshold Scenario ***********************************************************************
    
    
    if run == 10 and scenario == "80%" and filename == "shippingservice" and replicas == 2:
        filename = "frontend"
        run = 1
        scenario = "20%"
        replicas = 5
        CPU_Request = 100

    if run == 10 and scenario == "80%" and filename == "shippingservice" and replicas == 5:
        filename = "frontend"
        run = 1
        scenario = "20%"
        replicas = 10
        CPU_Request = 100

    run = run + 1




